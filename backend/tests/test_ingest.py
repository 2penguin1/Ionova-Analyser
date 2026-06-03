"""Workbook sheet-parsing logic (pure, no DB) — builds an in-memory .xlsx."""

import io

import openpyxl

from app.core.fields import SWIFT_FIELDS
from app.ingest.excel_importer import (
    _extract_address,
    _iter_result_rows,
    _parse_field_metrics,
    _parse_summary,
)


def _wb():
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary.append(["Run ID", "test-run-1"])
    summary.append(["Dataset Name", "TEST-DATA"])
    summary.append(["Overall F1", "0.9001"])

    fm = wb.create_sheet("Field Metrics")
    fm.append(["Field Name", "Correct", "Extra", "Missing", "Wrong", "Total"])
    fm.append(["Ctry", 10, 0, 1, 2, 13])

    results = wb.create_sheet("Results")
    header = ["Result ID", "Entry ID", "Status", "Execution Time (ms)", "Address Hash", "Input Address (truncated)"]
    for f in SWIFT_FIELDS:
        header += [f"Gold: {f}", f"Algo: {f}", f"Verdict: {f}"]
    results.append(header)
    row = ["r1", "e1", "FAILED", 12.5, "hash1", "{'Address': 'ACME GMBH | FRANKFURT, 60311, DE-HE'}"]
    for f in SWIFT_FIELDS:
        if f == "Ctry":
            row += ["DE", "DE", "Correct"]
        elif f == "TwnNm":
            row += ["Frankfurt am Main", "FRANKFURT", "Wrong"]
        else:
            row += [None, None, "Empty"]
    results.append(row)
    return wb


def test_parse_summary():
    s = _parse_summary(_wb()["Summary"])
    assert s["Run ID"] == "test-run-1"
    assert s["Overall F1"] == "0.9001"


def test_parse_field_metrics():
    rows = _parse_field_metrics(_wb()["Field Metrics"])
    assert rows == [{"field_name": "Ctry", "correct": 10, "extra": 0, "missing": 1, "wrong": 2, "total": 13}]


def test_extract_address():
    clean, raw = _extract_address("{'Address': 'A | B | C'}")
    assert clean == "A | B | C"
    assert raw == "{'Address': 'A | B | C'}"
    # Non-dict input passes through.
    assert _extract_address("plain text")[0] == "plain text"


def test_iter_result_rows_maps_triples_and_rollups():
    rows = list(_iter_result_rows(_wb()["Results"], "run-uuid"))
    assert len(rows) == 1
    r = rows[0]
    assert r["status"] == "FAILED"
    assert r["input_address"] == "ACME GMBH | FRANKFURT, 60311, DE-HE"
    assert r["gold_ctry"] == "DE" and r["algo_ctry"] == "DE" and r["verdict_ctry"] == "Correct"
    assert r["gold_twnnm"] == "Frankfurt am Main" and r["verdict_twnnm"] == "Wrong"
    # Rollups: 1 correct (Ctry), 1 wrong (TwnNm), rest Empty.
    assert r["n_correct"] == 1 and r["n_wrong"] == 1
    assert r["fields"]["TwnNm"]["verdict"] == "Wrong"
