"""Parse an exported eval-run workbook into normalized DB rows.

Built directly against the verified export format (4 sheets):
  - Summary       : (Metric, Value) key/value rows
  - Field Metrics : Field Name, Correct, Extra, Missing, Wrong, Total
  - Results       : base columns + Gold:/Algo:/Verdict: triple per field
  - Mismatches    : subset of Results (ignored — derivable from verdicts)

Import is idempotent on the source "Run ID": re-importing the same run replaces
the prior rows. Results are streamed and bulk-inserted in chunks so a 30k-row
workbook never fully materializes in memory.
"""

from __future__ import annotations

import ast as _ast
import io
from datetime import datetime
from typing import Any, Iterator

import openpyxl
from sqlalchemy import delete, insert
from sqlalchemy.orm import Session

from app.core.fields import SWIFT_FIELDS, algo_col, gold_col, verdict_col
from app.models import EvalResult, EvalRun, FieldMetric

CHUNK_SIZE = 2000

# Map verdict cell value -> rollup counter attribute on EvalResult.
_VERDICT_ROLLUP = {
    "Correct": "n_correct",
    "Wrong": "n_wrong",
    "Missing": "n_missing",
    "Extra": "n_extra",
}


class ImportError_(Exception):
    """Raised when a workbook cannot be parsed into the expected shape."""


# --------------------------------------------------------------------------- #
# Sheet parsers
# --------------------------------------------------------------------------- #

def _parse_summary(ws) -> dict[str, Any]:
    summary: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        key = row[0]
        value = row[1] if len(row) > 1 else None
        if key is None:
            continue
        summary[str(key).strip()] = value
    return summary


def _parse_field_metrics(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()

        def _int(idx: int) -> int:
            try:
                return int(row[idx]) if row[idx] is not None else 0
            except (TypeError, ValueError):
                return 0

        rows.append(
            {
                "field_name": name,
                "correct": _int(1),
                "extra": _int(2),
                "missing": _int(3),
                "wrong": _int(4),
                "total": _int(5),
            }
        )
    return rows


def _extract_address(raw: Any) -> tuple[str | None, str | None]:
    """Return (clean_address, raw_text). The Input Address cell is a stringified
    dict like ``{'Address': 'A | B | C'}``; pull the inner string, keep the raw."""
    if raw is None:
        return None, None
    raw_text = str(raw)
    try:
        parsed = _ast.literal_eval(raw_text)
        if isinstance(parsed, dict):
            # Prefer an 'Address' key; else first string value.
            val = parsed.get("Address")
            if val is None:
                for v in parsed.values():
                    if isinstance(v, str):
                        val = v
                        break
            return (str(val) if val is not None else raw_text), raw_text
    except (ValueError, SyntaxError):
        pass
    return raw_text, raw_text


def _result_column_map(header: tuple) -> dict[str, int]:
    """Map header label -> column index, tolerant of column reordering."""
    return {str(h).strip(): i for i, h in enumerate(header) if h is not None}


def _iter_result_rows(ws, run_id) -> Iterator[dict[str, Any]]:
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    col = _result_column_map(header)

    def get(row, label):
        idx = col.get(label)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in rows:
        if row is None or all(c is None for c in row):
            continue

        clean_addr, raw_addr = _extract_address(get(row, "Input Address (truncated)"))

        exec_ms = get(row, "Execution Time (ms)")
        try:
            exec_ms = float(exec_ms) if exec_ms is not None else None
        except (TypeError, ValueError):
            exec_ms = None

        record: dict[str, Any] = {
            "run_id": run_id,
            "source_result_id": _str_or_none(get(row, "Result ID")),
            "source_entry_id": _str_or_none(get(row, "Entry ID")),
            "status": _str_or_none(get(row, "Status")),
            "execution_time_ms": exec_ms,
            "address_hash": _str_or_none(get(row, "Address Hash")),
            "input_address": clean_addr,
            "input_address_raw": raw_addr,
            "n_correct": 0,
            "n_wrong": 0,
            "n_missing": 0,
            "n_extra": 0,
        }

        fields_blob: dict[str, dict[str, Any]] = {}
        for f in SWIFT_FIELDS:
            gold = _str_or_none(get(row, f"Gold: {f}"))
            algo = _str_or_none(get(row, f"Algo: {f}"))
            verdict = _str_or_none(get(row, f"Verdict: {f}"))
            record[gold_col(f)] = gold
            record[algo_col(f)] = algo
            record[verdict_col(f)] = verdict
            fields_blob[f] = {"gold": gold, "algo": algo, "verdict": verdict}
            rollup = _VERDICT_ROLLUP.get(verdict or "")
            if rollup:
                record[rollup] += 1

        record["fields"] = fields_blob
        yield record


def _str_or_none(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v)
    return s if s != "" else None


def _parse_dt(v: Any) -> datetime | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    try:
        return datetime.fromisoformat(str(v))
    except ValueError:
        return None


# --------------------------------------------------------------------------- #
# Orchestrator
# --------------------------------------------------------------------------- #

def import_workbook(
    db: Session,
    file_bytes: bytes,
    *,
    import_batch_id=None,
) -> dict[str, Any]:
    """Parse + persist a workbook. Returns counts. Idempotent on source Run ID."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    names = set(wb.sheetnames)
    if "Summary" not in names or "Results" not in names:
        raise ImportError_(
            f"Workbook missing required sheets. Found: {sorted(names)}"
        )

    summary = _parse_summary(wb["Summary"])
    source_run_id = _str_or_none(summary.get("Run ID"))
    if not source_run_id:
        raise ImportError_("Summary sheet has no 'Run ID'.")

    # Idempotency: replace any prior import of the same run.
    existing = db.query(EvalRun).filter_by(source_run_id=source_run_id).first()
    if existing:
        db.delete(existing)
        db.flush()

    run = EvalRun(
        source_run_id=source_run_id,
        import_batch_id=import_batch_id,
        dataset_name=_str_or_none(summary.get("Dataset Name")),
        run_type=_str_or_none(summary.get("Run Type")),
        status=_str_or_none(summary.get("Status")),
        initiated_by=_str_or_none(summary.get("Initiated By")),
        started_at=_parse_dt(summary.get("Started At")),
        completed_at=_parse_dt(summary.get("Completed At")),
        algorithm_version=_str_or_none(summary.get("Algorithm Version")),
        summary=summary,
    )
    db.add(run)
    db.flush()  # assign run.id

    # Field metrics
    fm_count = 0
    if "Field Metrics" in names:
        for fm in _parse_field_metrics(wb["Field Metrics"]):
            db.add(FieldMetric(run_id=run.id, **fm))
            fm_count += 1

    # Results — chunked bulk insert
    result_count = 0
    buffer: list[dict[str, Any]] = []
    for record in _iter_result_rows(wb["Results"], run.id):
        buffer.append(record)
        if len(buffer) >= CHUNK_SIZE:
            db.execute(insert(EvalResult), buffer)
            result_count += len(buffer)
            buffer.clear()
    if buffer:
        db.execute(insert(EvalResult), buffer)
        result_count += len(buffer)

    counts = {
        "run_id": str(run.id),
        "source_run_id": source_run_id,
        "field_metrics": fm_count,
        "results": result_count,
    }
    run.summary = {**summary, "_import_counts": counts}
    db.flush()
    wb.close()
    return counts
